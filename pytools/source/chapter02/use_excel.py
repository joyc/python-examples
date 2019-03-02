from datetime import datetime
import MySQLdb

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors


class ExcelUtils(object):
    """
    pip install openpyxl
    pip install pillow
    """
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_two = self.wb.create_sheet('我的表单')
        self.ws.title = '你的表单'
        self.ws.sheet_properties.tabColor = 'ff0000'
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        # 插入数据
        self.ws['A1'] = 66
        self.ws['A2'] = '你好'
        self.ws['A3'] = datetime.now()

        for row in self.ws_two['A1:E5']:
            for cell in row:
                cell.value = 2

        # 对数据进行求和
        self.ws_two['G1'] = '=SUM(A1:E1)'

        # 设置文字
        font = Font(sz=18, color=colors.RED)
        self.ws['A2'].font = font

        # 插入图片
        # img = Image('./static/temp.jpg')
        # self.ws.add_image(img, 'B1')

        self.ws.merge_cells('A4:E5')
        self.ws.unmerge_cells('A4:E5')
        self.wb.save('./static/test.xlsx')

    def read_xls(self):
        """
         读取excel数据
        :return:
        """
        ws = load_workbook('./static/template.xlsx')
        names = ws.get_sheet_names()
        print(names)

        conn = self.get_conn()
        wb = ws.active
        wb = ws[names[0]]
        for (i, row) in enumerate(wb.rows):
            if i < 2:
                continue
            year = wb['A{0}'.format(i + 1)].value
            max = wb['B{0}'.format(i + 1)].value
            avg = wb['C{0}'.format(i + 1)].value
            print(year)
            if year is None:
                continue
            cursor = conn.cursor()
            sql = 'INSERT INTO `score`(`year`, `max`, `avg`) VALUES({year}, {max}, {avg})'.format(
                year=year, max=max, avg=avg)
            print(sql)
            cursor.execute(sql)
            conn.autocommit(True)
            # print(conn)

    def get_conn(self):
        """ 获取mysql 的连接 """
        try:
            conn = MySQLdb.connect(
                db='user_grade',
                host='localhost',
                user='root',
                password='',
                charset='utf8'
            )
        except:
            pass
        return conn

    def export_xls(self):
        """ 从mysql数据库导出数据到excel """
        # 获取数据库的连接
        conn = self.get_conn()
        cursor = conn.cursor()
        # 准备查询语句 (如果数据量大，需要借助于分页查询)
        sql = 'SELECT `year`, `max`, `avg` FROM `score`'
        # 查询数据
        cursor.execute(sql)
        rows = cursor.fetchall()

        # 循环写入到excel
        wb = Workbook()
        ws = wb.active
        for (i, row) in enumerate(rows):
            print(row)
            (ws['A{0}'.format(i+1)],
            ws['B{0}'.format(i+1)],
            ws['C{0}'.format(i+1)]) = row

        # 保存excel
        wb.save('./static/export.xlsx')


if __name__ == '__main__':
    client = ExcelUtils()
    # client.do_sth()
    client.read_xls()
    # client.export_xls()
